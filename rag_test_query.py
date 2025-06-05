# rag_query_rfp.py
from openpyxl import load_workbook
from langchain.vectorstores import FAISS
from langchain.embeddings import OpenAIEmbeddings

# Load RFP questions (non-bold, column B) from sheet
wb = load_workbook("RFP-Document.xlsx")
ws = wb["RFP Questions"]
questions = []
for row in ws.iter_rows(min_row=0):
    cell = row[1]
    if cell.value and not cell.font.bold:
        questions.append(cell.value.strip())

# Load FAISS vectorstore correctly
embedding = OpenAIEmbeddings(model="text-embedding-ada-002")
vectorstore = FAISS.load_local(
    "faiss_index",
    embeddings=embedding,
    allow_dangerous_deserialization=True,  # ✅ Trust your own saved file
)


# Query each question
for idx, question in enumerate(questions):
    print(f"\n🔍 Q{idx+1}: {question}")
    try:
        docs = vectorstore.similarity_search(question, k=3)
        for i, doc in enumerate(docs):
            print(f"   ➤ Context {i+1}:\n{doc.page_content.strip()[:500]}...\n")
    except Exception as e:
        print(f"   ❌ Error: {e}")
