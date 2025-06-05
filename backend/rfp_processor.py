import os
import tempfile
import pandas as pd
import re
import numpy as np
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from langchain_community.vectorstores import FAISS
from langchain_openai import OpenAIEmbeddings
from openai import OpenAI
from backend.utils import read_file


class RFPProcessor:
    def __init__(self):
        self.embedding = OpenAIEmbeddings(model="text-embedding-ada-002")
        self.openai_client = OpenAI()
        self.training_qa_pairs = {}  # Store training Q&A pairs
        self.training_embeddings = {}  # Cache for training question embeddings

    def extract_questions_from_excel(self, file_path: str) -> List[str]:
        """Extract questions from Excel using the proven logic, skipping instruction sheets and date-like entries"""
        import re
        from datetime import datetime

        def is_date(text: str) -> bool:
            # Common date patterns
            date_patterns = [
                r"\d{4}-\d{2}-\d{2}",  # YYYY-MM-DD
                r"\d{2}/\d{2}/\d{4}",  # MM/DD/YYYY
                r"[A-Za-z]+ \d{1,2},? \d{4}",  # Month DD, YYYY
                r"\d{1,2} [A-Za-z]+ \d{4}",  # DD Month YYYY
            ]

            # Check if text matches any date pattern
            if any(re.search(pattern, text) for pattern in date_patterns):
                return True

            # Try parsing as datetime
            try:
                # Remove time if present
                text = text.split(" ", 1)[0] if " " in text else text
                datetime.strptime(text, "%Y-%m-%d")
                return True
            except ValueError:
                return False

        wb = load_workbook(file_path)
        all_questions = []

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip instruction sheets
            first_cells_text = " ".join(
                [
                    str(cell.value or "").lower()
                    for row in ws.iter_rows(max_row=3)
                    for cell in row
                    if cell.value
                ]
            )
            if "instruction" in first_cells_text or "français" in first_cells_text:
                continue

            # Extract questions from current sheet
            for row in ws.iter_rows(min_row=1):
                q_cell = row[1] if len(row) > 1 else None  # Column B
                if q_cell and q_cell.value and not q_cell.font.bold:
                    question = str(q_cell.value).strip()
                    # Only add if it's not a date and has proper sentence structure
                    if (
                        not is_date(question) and len(question.split()) > 3
                    ):  # Basic check for sentence-like structure
                        all_questions.append(question)

        return all_questions

    def load_training_qa_pairs(self, excel_path: str):
        """Load Q&A pairs from training Excel file and precompute embeddings"""
        wb = load_workbook(excel_path)

        new_questions = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip instruction sheets
            first_cells_text = " ".join(
                [
                    str(cell.value or "").lower()
                    for row in ws.iter_rows(max_row=3)
                    for cell in row
                    if cell.value
                ]
            )
            if "instruction" in first_cells_text or "français" in first_cells_text:
                continue

            # Extract Q&A pairs from current sheet
            for row in ws.iter_rows(min_row=1):
                q_cell = row[1] if len(row) > 1 else None
                r_cell = row[2] if len(row) > 2 else None

                if (
                    q_cell
                    and q_cell.value
                    and not q_cell.font.bold
                    and r_cell
                    and r_cell.value
                ):
                    question = str(q_cell.value).strip()
                    if question not in self.training_qa_pairs:
                        new_questions.append(question)
                        self.training_qa_pairs[question] = str(r_cell.value).strip()

        # Batch compute embeddings for new questions
        if new_questions:
            embeddings = self.embedding.embed_documents(new_questions)
            for q, emb in zip(new_questions, embeddings):
                self.training_embeddings[q] = emb

    def calculate_similarity(self, q1: str, q2: str) -> float:
        """Calculate semantic similarity between two questions"""
        try:
            emb1 = self.embedding.embed_query(q1)
            emb2 = self.embedding.embed_query(q2)
            return np.dot(emb1, emb2) / (np.linalg.norm(emb1) * np.linalg.norm(emb2))
        except Exception:
            return 0.0

    def find_similar_questions(
        self, question: str, threshold: float = 0.85
    ) -> List[Tuple[str, float]]:
        """Find similar questions using batch embedding comparison"""
        if not self.training_embeddings:
            return []

        # Get embedding for the current question
        q_embedding = self.embedding.embed_query(question)

        # Calculate similarities in bulk using numpy
        similarities = []
        for train_q, train_emb in self.training_embeddings.items():
            similarity = np.dot(q_embedding, train_emb) / (
                np.linalg.norm(q_embedding) * np.linalg.norm(train_emb)
            )
            if similarity >= threshold:
                similarities.append((train_q, similarity))

        return sorted(similarities, key=lambda x: x[1], reverse=True)

    def evaluate_context_sufficiency(self, question: str, context: str) -> bool:
        """Check if context is sufficient to answer the question"""
        if not context:
            return False

        prompt = f"""Given this context: '{context}'
        
        Can you provide a complete and accurate answer to this question: '{question}'?
        
        Rate the context sufficiency on a scale of 0-100 where:
        0-84: Insufficient context
        85-100: Sufficient context
        
        Reply with just the number."""

        response = self.openai_client.chat.completions.create(
            model="gpt-4",
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert at evaluating context sufficiency for answering questions.",
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=10,
        )

        try:
            score = int(response.choices[0].message.content.strip())
            return score >= 85
        except:
            return False

    def process_rfp(self, file, vectorstore) -> Tuple[Dict, List[str]]:
        """Process RFP and return answers and questions needing more context"""
        # Create a temporary file to process
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file.getvalue())
            tmp_path = tmp.name

            try:
                # Extract questions
                questions = self.extract_questions_from_excel(tmp_path)
            finally:
                try:
                    os.remove(tmp_path)
                except:
                    pass

        answers = {}
        needs_context = []

        # Process questions in batches
        batch_size = 5  # Process 5 questions at a time
        for i in range(0, len(questions), batch_size):
            batch = questions[i : i + batch_size]

            for question in batch:
                # Find similar questions efficiently
                similar_questions = self.find_similar_questions(question)

                if similar_questions:
                    # Use best matching answer
                    best_match = similar_questions[0][0]
                    answers[question] = self.training_qa_pairs[best_match]
                    continue

                # If no match found, use RAG
                docs = vectorstore.similarity_search(question, k=3)
                context = "\n\n".join([doc.page_content for doc in docs])

                # Check if we have sufficient context
                if not self.evaluate_context_sufficiency(question, context):
                    needs_context.append(question)
                    continue

                # Generate answer using context
                response = self.openai_client.chat.completions.create(
                    model="gpt-4",
                    messages=[
                        {
                            "role": "system",
                            "content": "Generate a precise answer based on the provided context.",
                        },
                        {
                            "role": "user",
                            "content": f"Context:\n{context}\n\nQuestion: {question}",
                        },
                    ],
                    temperature=0.2,
                )

                answer = response.choices[0].message.content
                answers[question] = answer

        return answers, needs_context
