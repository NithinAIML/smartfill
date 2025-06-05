import os
import pandas as pd
import numpy as np
import fitz  # PyMuPDF
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
from backend.utils import clean_text
from backend.config import settings
from openpyxl import load_workbook
import tempfile
from docx import Document


class RAGEngine:
    def __init__(self, layout_aware=False):
        self.embedding = OpenAIEmbeddings(model="text-embedding-ada-002")
        # Initialize FAISS with a dummy document to avoid empty index error
        dummy_text = "initialization document"
        self.db = FAISS.from_texts([dummy_text], self.embedding)
        # Clear the dummy document
        self.db._index = None
        self.layout_aware = layout_aware
        self.splitter = RecursiveCharacterTextSplitter(
            chunk_size=500, chunk_overlap=50, separators=["\n\n", "\n", ".", " "]
        )

    def process_pdf(self, file_path):
        """Special handling for PDF documents - page by page extraction"""
        chunks = []
        with fitz.open(file_path) as doc:
            for page_num, page in enumerate(doc):
                text = page.get_text("text")
                if text.strip():
                    # Add page metadata to help with context
                    page_text = f"[Page {page_num + 1}]\n{text}"
                    # Chunk the page text
                    page_chunks = self.splitter.split_text(page_text)
                    chunks.extend(page_chunks)

        # Add chunks to vector store with page metadata
        if chunks:
            self.db.add_texts(chunks)
        return len(chunks)

    def process_excel_qa(self, file_path):
        """Special handling for Excel Q&A format"""
        wb = load_workbook(file_path)

        qa_chunks = []
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip instruction sheets by checking first few cells for instruction-like content
            first_cells_text = " ".join(
                [
                    str(cell.value or "").lower()
                    for row in ws.iter_rows(max_row=3)
                    for cell in row
                    if cell.value
                ]
            )
            if "instruction" in first_cells_text or "français" in first_cells_text:
                continue

            questions = []
            responses = []
            pending_question = None

            # Extract Q&A pairs using the proven logic
            for row in ws.iter_rows(min_row=1):
                q_cell = row[1] if len(row) > 1 else None  # Column B
                r_cell = row[2] if len(row) > 2 else None  # Column C

                if q_cell and q_cell.value and not q_cell.font.bold:
                    pending_question = str(q_cell.value).strip()

                    if r_cell and r_cell.value:
                        questions.append(pending_question)
                        responses.append(str(r_cell.value).strip())
                        pending_question = None

                elif pending_question and r_cell and r_cell.value:
                    questions.append(pending_question)
                    responses.append(str(r_cell.value).strip())
                    pending_question = None

            # Create specialized chunks that preserve Q&A mapping
            sheet_chunks = [
                f"Question: {q}\nAnswer: {a}" for q, a in zip(questions, responses)
            ]
            qa_chunks.extend(sheet_chunks)

        # Add to vector store with Q&A metadata
        if qa_chunks:
            self.db.add_texts(qa_chunks)
        return len(qa_chunks)

    def add_document(self, file):
        """Main entry point for adding documents"""
        ext = os.path.splitext(file.name)[1].lower()

        # Create a temporary file to process
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(file.getvalue())  # Use getvalue() for Streamlit's UploadedFile
            tmp_path = tmp.name

            try:
                # Process based on file type
                if ext == ".pdf":
                    chunks = []
                    with fitz.open(tmp_path) as doc:
                        for page_num, page in enumerate(doc):
                            text = page.get_text("text")
                            if text.strip():
                                page_text = f"[Page {page_num + 1}]\n{text}"
                                page_chunks = self.splitter.split_text(page_text)
                                chunks.extend(page_chunks)

                elif ext == ".xlsx":
                    wb = load_workbook(tmp_path)
                    chunks = []

                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        first_cells_text = " ".join(
                            [
                                str(cell.value or "").lower()
                                for row in ws.iter_rows(max_row=3)
                                for cell in row
                                if cell.value
                            ]
                        )
                        if (
                            "instruction" in first_cells_text
                            or "français" in first_cells_text
                        ):
                            continue

                        text = "\n".join(
                            str(cell.value).strip()
                            for row in ws.iter_rows()
                            for cell in row
                            if cell.value
                        )
                        sheet_chunks = self.splitter.split_text(text)
                        chunks.extend(sheet_chunks)

                elif ext == ".docx":
                    doc = Document(tmp_path)
                    text = "\n".join(
                        [para.text for para in doc.paragraphs if para.text.strip()]
                    )
                    chunks = self.splitter.split_text(text)

                # Add chunks to vector store
                if chunks:
                    self.db.add_texts(chunks)
                    return len(chunks)
                return 0

            finally:
                # Clean up temporary file
                try:
                    os.remove(tmp_path)
                except:
                    pass

    def retrieve(self, question, k=4):
        """Enhanced retrieval with metadata handling"""
        docs = self.db.similarity_search(question, k=k)
        results = []

        for doc in docs:
            content = doc.page_content
            # Extract page number if present
            if "[Page " in content:
                page_info = content.split("]")[0] + "]"
                content = content.replace(page_info, "").strip()
            results.append(content)

        return results if results else []

    def answer(self, question):
        context = self.retrieve(question)
        if not context:
            return "No relevant information found."

        prompt = f"Context:\n{''.join(context)}\n\nQuestion: {question}\nAnswer:"

        from openai import OpenAI

        client = OpenAI()
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {
                    "role": "system",
                    "content": "You are a helpful assistant providing accurate answers based on the given context.",
                },
                {"role": "user", "content": prompt},
            ],
        )
        return response.choices[0].message.content.strip()
