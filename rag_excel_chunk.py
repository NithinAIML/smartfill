from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings
from langchain.vectorstores import FAISS
import pandas as pd

from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("Attachment A.xlsx")
ws = wb.active

questions = []
responses = []
pending_question = None

for row in ws.iter_rows(min_row=1):
    q_cell = row[1] if len(row) > 1 else None  # Column B
    r_cell = row[2] if len(row) > 2 else None  # Column C

    if q_cell and q_cell.value and not q_cell.font.bold:
        pending_question = q_cell.value.strip()

        # If response exists in same row (col C)
        if r_cell and r_cell.value:
            questions.append(pending_question)
            responses.append(str(r_cell.value).strip())
            pending_question = None

    elif pending_question and r_cell and r_cell.value:
        # Handle if response comes later
        questions.append(pending_question)
        responses.append(str(r_cell.value).strip())
        pending_question = None

# Final cleanup
qa_df = pd.DataFrame({"question": questions, "response": responses})
print(f"✅ Extracted {len(qa_df)} Q&A pairs")
print(qa_df.head(25))

# Load the Q&A Excel data you extracted
# qa_df = pd.read_excel("Attachment A.xlsx")
# print(qa_df.columns.tolist())

# Combine Question + Response
# qa_pairs = [f"Q: {q}\nA: {a}" for q, a in zip(qa_df["RFP Questions"], qa_df["Vendor Response"])]

# # Chunking
# splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=50)
# chunks = splitter.split_text("\n\n".join(qa_pairs))
# print(chunks)

# # Embedding
# embedding = OpenAIEmbeddings(model="text-embedding-ada-002")
# vectorstore = FAISS.from_texts(chunks, embedding)

# # Save to FAISS index folder
# vectorstore.save_local("faiss_index")
# print(f"✅ FAISS index built and saved with {len(chunks)} chunks")
